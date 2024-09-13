from django.http import HttpResponse
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import permission_required
from django.http import HttpResponse
from django.db import IntegrityError
from .forms import ClientForm , AdherenteForm, DateRangeForm
from django.contrib import messages
from .models import Titular , Adherente, Log
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook

def home(request):
    return render(request, 'signin.html')

def signup(request):
    if request.method == 'GET':
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(username=request.POST['username'],
                                                password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('create_client')
            except IntegrityError:
                return render(request, 'signup.html', {
                    'form': UserCreationForm,
                    "error": 'Username already exists'
                })

        return render(request, 'signup.html', {
            'form': UserCreationForm,
            "error": 'Password do not match'
        })

@login_required
def client(request):
    titulares = Titular.objects.filter(user_upload=request.user)
    return render(request, 'clients.html',{'titulares': titulares})


def signout(request):
    logout(request)
    return redirect('home')


def login_user(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {'form': AuthenticationForm})
    else:
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'signin.html', {'form': AuthenticationForm, 'error': 'Usuario o pasword incorrecto'})
        else:
            login(request, user)
            return redirect('select_cbu')

@login_required
def create_client(request):
    if request.method == 'GET':
        return render(request, 'create_client.html', {
            'form': ClientForm()
        })
    elif request.method == 'POST':
        try:
            form = ClientForm(request.POST)
            if form.is_valid():
                new_client = form.save(commit=False)
                new_client.user_upload = request.user
                new_client.save()
                return render(request, 'create_client.html', {
                    'new_client': new_client,
                    'form2': AdherenteForm()
                })
            else:
                # Mostrar los errores específicos del formulario
                print(form.errors)
                return render(request, 'create_client.html', {
                    'form': form,
                    'error': form.errors  # Muestra los errores del formulario
                })
        except ValueError as e:
            print(f"ValueError: {e}")
            return render(request, 'create_client.html', {
                'form': ClientForm(),
                'error': 'Please provide valid data'
            })
        except Exception as e:
            print(f"Unexpected error: {e}")
            return render(request, 'create_client.html', {
                'form': ClientForm(),
                'error': 'An unexpected error occurred'
            })


@login_required
def create_adherente(request):
    if request.method == 'POST':
        try:
            form = AdherenteForm(request.POST)
            if form.is_valid():
                titular_id = request.POST.get('titular_id')
                
                titular = Titular.objects.get(pk=titular_id)
                new_adherente = form.save(commit=False)
                new_adherente.user_upload = request.user
                new_adherente.titular = titular
                new_adherente.save()

                Log.objects.create(
                    adherente=new_adherente,
                    movimiento='Creacion',
                    user=request.user
                )

                titular_id = request.POST.get('titular_id')
                new_client = Titular.objects.get(titular_id=titular_id)
                adherentes = Adherente.objects.filter(titular=titular_id)
                return render(request, 'create_client.html', {
                'new_client': new_client,
                'form2': form, # Redirigir a una vista de listado de clientes o a donde sea apropiado
                'tupla_adherentes': adherentes
                })
            else:
                raise ValueError
        except ValueError:
            titular_id = request.POST.get('titular_id')
            new_client = Titular.objects.get(titular_id=titular_id)
            return render(request, 'create_client.html', {
                'new_client': new_client,
                'form2': form,
                'error': 'Please provide valid data for adherente'
            })
    elif request.method == 'GET':
        titular_id = request.GET.get('titular_id')
        print(titular_id)
        new_client = Titular.objects.get(titular_id=titular_id)
        adherentes = Adherente.objects.filter(titular=titular_id)
        return render(request, 'create_adherentes.html', {
                'new_client': new_client,
                'tupla_adherentes': adherentes
                })



@login_required
def client_detail(request, titular_id):
    if request.method == 'GET':
        #cliente = Titular.objects.filter(titular_id=titular_id)
        #titular = Titular.objects.get(pk=titular_id) esto puede tumbar el servidor
        titular = get_object_or_404(Titular,pk=titular_id)
        form = ClientForm(instance=titular)
        return render(request, 'client_detail.html',{'titular': titular, 'form': form})
    elif request.method == 'POST':
        try:
            titular = get_object_or_404(Titular,pk=titular_id)
            #titular = get_object_or_404(Titular,pk=titular_id,user_upload=request.user )
            form = ClientForm(request.POST, instance=titular)
            form.save()
            return redirect('clients')
        except ValueError:
            return render(request, 'client_detail.html',{'titular': titular, 'form': form, 'error': 'error actualizando cliente'})


@login_required
def client_baja(request,titular_id):
    titular = get_object_or_404(Titular, pk=titular_id)
    if request.method == 'POST':
        titular.is_active = False
        titular.deleted = timezone.now()
        titular.save()
        adherentes = Adherente.objects.filter(titular=titular_id)
        return render(request, 'create_client.html', {
            'new_client': titular,
            'tupla_adherentes': adherentes
        })

@login_required
def consultar_cbu(request):
    if request.method == 'GET':
        return render(request, 'create_client_selcb.html', {
        })
    elif request.method == 'POST':
        cbu_r = request.POST.get("cbu")
        try:
            new_client = Titular.objects.get(cbu=cbu_r)
            try:
                titular_id = new_client.titular_id
                adherentes = Adherente.objects.filter(titular=titular_id)
                return render(request, 'create_client.html', {
                'new_client': new_client,
                'tupla_adherentes': adherentes,
                })
            except:
                return render(request, 'create_client.html', {
                'new_client': new_client,
                'cbu': cbu_r,
                })
        except Titular.DoesNotExist:
                return render(request, 'create_client.html', {
                'form': ClientForm(),
                'cbu': cbu_r,
            })

@login_required
def bajaAdherente(request, adherente_id):
    adherente = Adherente.objects.get(pk=adherente_id)
    if request.method == 'POST':
        adherente.is_active = False
        adherente.save()

        titular_id = adherente.titular.pk  # Obtener el ID del titular correctamente
        new_client = get_object_or_404(Titular, pk=titular_id)
        adherentes = Adherente.objects.filter(titular=titular_id)

        Log.objects.create(
                    adherente=adherente,
                    movimiento='Baja',
                    user=request.user
                )
        
        return render(request, 'create_client.html', {
            'new_client': new_client,
            'tupla_adherentes': adherentes
        })
    else:
        return HttpResponse("none")


@login_required
def updateAdherente(request, adherente_id):
    adherente = Adherente.objects.get(pk=adherente_id)
    checklist = adherente.is_active
    if request.method == 'POST':
        form = AdherenteForm(request.POST, instance=adherente)
        if form.is_valid():
            form.save()

            titular_id = adherente.titular.titular_id
            adherentes = Adherente.objects.filter(titular=titular_id)

            Log.objects.create(
                    adherente=adherente,
                    movimiento='Modificacion',
                    user=request.user
                )
            
            new_client = Titular.objects.get(pk=titular_id)
            return render(request, 'create_client.html', {
            'new_client': new_client,
            'tupla_adherentes': adherentes
            })
    else:
        form = AdherenteForm(instance=adherente)

    return render(request, 'update_adherente.html', {'adherente': adherente, 'form': form, 'is_active': checklist})

@login_required
def update_titular(request, titular_id):
    titular = get_object_or_404(Titular, pk=titular_id)
    checklist = titular.is_active
    print(titular.__dict__) 
    form = ClientForm(request.POST, instance=titular)

    if request.method == 'POST':

        if form.is_valid():
            
            form.save()
            adherentes = Adherente.objects.filter(titular=titular)
            return render(request, 'create_client.html', {
            'new_client': titular,
            'tupla_adherentes': adherentes
            }) # Redirigir después de guardar
        else:
            form = ClientForm(instance=titular)
            print(titular.between_street)
            print("Formulario inválido", form.errors)

    elif request.method == 'GET':
        pass
    return render(request, 'update_client.html', {'client': titular, 'form': form, 'is_active' : checklist})

@login_required
def buscar(request):
    if request.method == 'GET':
        return render(request, 'buscar.html')
    elif request.method == 'POST':
        nombre = request.POST.get('name')
        telefono = request.POST.get('phone')
        direccion = request.POST.get('address')
        dni = request.POST.get('dni')
        activo = request.POST.get('is_active')
        cbu = request.POST.get('cbu')


        if not any([cbu, nombre, telefono, direccion, dni]):
                    return render(request, 'buscar.html', {
                        'error': 'Debe proporcionar al menos un criterio de búsqueda.'
                    })

        filtro = {}

        if cbu:
            filtro['cbu'] = cbu
        if nombre:
            filtro['name__icontains'] = nombre
        if telefono:
            filtro['phone'] = telefono
        if direccion:
            filtro['address__icontains'] = direccion
        if dni:
            filtro['dni'] = dni
        if activo:
            filtro['is_active'] = True


        titulares = Titular.objects.filter(**filtro)
        if titulares.exists():

            return render(request, 'list_client.html', {
                'new_client': titulares
            })
        else:
            # Manejar el caso en que no se encuentren titulares
            return render(request, 'list_client.html', {
                'new_client': None,
                'error': 'No se encontraron titulares con los datos proporcionados.'
            })



@permission_required('pianoadherentes.can_view_stats', raise_exception=True)
def stadisticas(request):


    return render(request, 'estadisticas.html')


@permission_required('pianoadherentes.can_view_stats', raise_exception=True)
def generate_excel(request):
    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')

            # Convertir las fechas de cadena a objetos datetime
            start_date_parts = start_date_str.split('/')
            start_date = '-'.join(reversed(start_date_parts))

            end_date_parts = end_date_str.split('/')
            end_date = '-'.join(reversed(end_date_parts))

            # Filtrar adherentes dentro del rango de fechas
            adherentes = Adherente.objects.filter(adherente_date__range=(start_date, end_date))

            # Crear un archivo Excel utilizando openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = 'Adherentes'

            # Definir encabezados de columnas
            headers = [
                'Adherente ID',
                'Titular',
                'DNI Titular',
                'CBU Titular',
                'Nombre',
                'Teléfono',
                'Dirección',
                'DNI',
                'Fecha de Creación',
                'Usuario',
            ]

            # Escribir encabezados en la primera fila
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num, value=header)

            # Escribir datos de adherentes en filas
            for row_num, adherente in enumerate(adherentes, start=2):
                ws.cell(row=row_num, column=1, value=adherente.adherente_id)
                ws.cell(row=row_num, column=2, value=adherente.titular.name)
                ws.cell(row=row_num, column=3, value=adherente.titular.dni)
                ws.cell(row=row_num, column=4, value=adherente.titular.cbu)
                ws.cell(row=row_num, column=5, value=adherente.name)
                ws.cell(row=row_num, column=6, value=adherente.phone)
                ws.cell(row=row_num, column=7, value=adherente.address)
                ws.cell(row=row_num, column=8, value=adherente.dni)
                ws.cell(row=row_num, column=9, value=adherente.adherente_date.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=10, value=adherente.user_upload.username)

            # Crear el archivo Excel en memoria
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="informe.xlsx"'

            # Guardar el libro de trabajo en la respuesta HTTP
            wb.save(response)

            return response
    else:
        form = DateRangeForm()

    return render(request, 'estadisticas.html', {'form': form})