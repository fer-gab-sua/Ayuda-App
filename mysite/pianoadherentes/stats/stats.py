from django.contrib.auth.decorators import permission_required
from ..models import Titular , Adherente, Log, Sucursales
from django.contrib.auth.models import User
from openpyxl import Workbook
from django.http import HttpResponse
from django.shortcuts import render

@permission_required('pianoadherentes.can_view_stats', raise_exception=True)
def estadisticas(request):
    sucursales = Sucursales.objects.all()
    usuarios = User.objects.all()
    return render(request, 'estadisticas.html',{"sucursales":sucursales, "usuarios":usuarios}) 





@permission_required('pianoadherentes.can_view_stats', raise_exception=True)
def mis_ventas(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('init_date')
        end_date_str = request.POST.get('end_date')

        print(request.POST.get('init_date'))
        print(end_date_str)

        # Convertir las fechas de cadena a objetos datetime
        start_date_parts = start_date_str.split('-')
        start_date = '-'.join(start_date_parts)

        end_date_parts = end_date_str.split('-')
        end_date = '-'.join(end_date_parts)

        # Filtrar adherentes dentro del rango de fechas
        adherentes = Adherente.objects.filter(
            created__range=(start_date, end_date),
            user_upload=request.user)

        # Crear un archivo Excel utilizando openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = 'Adherentes'

        # Definir encabezados de columnas
        headers = [
            'Adherente ID',
            'Titular Nombre',
            'Titular Apellido',
            'Tipo de documento',
            'Documento',
            'CBU Titular',
            'Nombre',
            'Apellido',
            'Teléfono',
            'Dirección',
            'Tipo documento',
            'Nro documento',
            'Fecha de Creación',
            'Usuario',
            'Sucursal'
        ]

        # Escribir encabezados en la primera fila
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos de adherentes en filas
        for row_num, adherente in enumerate(adherentes, start=2):
            ws.cell(row=row_num, column=1, value=adherente.adherente_id)
            ws.cell(row=row_num, column=2, value=adherente.titular.name)
            ws.cell(row=row_num, column=3, value=adherente.titular.last_name)
            ws.cell(row=row_num, column=4, value=adherente.titular.document_type)
            ws.cell(row=row_num, column=5, value=adherente.titular.document)
            ws.cell(row=row_num, column=6, value=adherente.titular.cbu)
            ws.cell(row=row_num, column=7, value=adherente.name)
            ws.cell(row=row_num, column=8, value=adherente.last_name)
            ws.cell(row=row_num, column=9, value=adherente.phone)
            address_complete = str(f"{adherente.street_address} {adherente.number} {adherente.floor} - {adherente.city} - {adherente.province}")
            ws.cell(row=row_num, column=10, value=address_complete)
            ws.cell(row=row_num, column=11, value=adherente.document_type)
            ws.cell(row=row_num, column=12, value=adherente.document)
            ws.cell(row=row_num, column=13, value=adherente.created.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=14, value=adherente.user_upload.username)
            ws.cell(row=row_num, column=15, value=adherente.sucursal)

        # Crear el archivo Excel en memoria
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="informe.xlsx"'
        # Guardar el libro de trabajo en la respuesta HTTP
        wb.save(response)
        return response
    return render(request, 'estadisticas.html')




@permission_required('pianoadherentes.can_view_stats', raise_exception=True)
def mis_log(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('init_date')
        end_date_str = request.POST.get('end_date')

        # Convertir las fechas de cadena a objetos datetime
        start_date_parts = start_date_str.split('-')
        start_date = '-'.join(start_date_parts)

        end_date_parts = end_date_str.split('-')
        end_date = '-'.join(end_date_parts)

        # Filtrar adherentes dentro del rango de fechas

        log = Log.objects.filter(
            created__range=(start_date, end_date),
            user=request.user)

        # Crear un archivo Excel utilizando openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = 'My_log'

        # Definir encabezados de columnas
        headers = [
            'Log ID',
            'Adherente',
            'Movimiento',
            'Usuario',
            'Historia',
            'Fecha'
        ]

        # Escribir encabezados en la primera fila
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos de adherentes en filas
        for row_num, log in enumerate(log, start=2):
            ws.cell(row=row_num, column=1, value=log.log_id)
            adherente = str(f'{log.adherente.name} {log.adherente.last_name}')
            ws.cell(row=row_num, column=2, value=adherente)
            ws.cell(row=row_num, column=3, value=log.movimiento)
            ws.cell(row=row_num, column=4, value=log.user.username)
            ws.cell(row=row_num, column=5, value=log.historia)
            ws.cell(row=row_num, column=6, value=log.created.strftime('%Y-%m-%d %H:%M:%S'))
            

        # Crear el archivo Excel en memoria
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="informe.xlsx"'
        # Guardar el libro de trabajo en la respuesta HTTP
        wb.save(response)
        return response
    return render(request, 'estadisticas.html')

def sucursal_ventas(request):
    if request.method == 'POST':
        print("llego hasta aca")
        start_date_str = request.POST.get('init_date')
        end_date_str = request.POST.get('end_date')
        sucursal = request.POST.get('sucursal')

        # Convertir las fechas de cadena a objetos datetime
        start_date_parts = start_date_str.split('-')
        start_date = '-'.join(start_date_parts)

        end_date_parts = end_date_str.split('-')
        end_date = '-'.join(end_date_parts)

        # Filtrar adherentes dentro del rango de fechas
        adherentes = Adherente.objects.filter(
            created__range=(start_date, end_date),
            sucursal=sucursal)

        # Crear un archivo Excel utilizando openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = 'Adherentes'

        # Definir encabezados de columnas
        headers = [
            'Adherente ID',
            'Titular Nombre',
            'Titular Apellido',
            'Tipo de documento',
            'Documento',
            'CBU Titular',
            'Nombre',
            'Apellido',
            'Teléfono',
            'Dirección',
            'Tipo documento',
            'Nro documento',
            'Fecha de Creación',
            'Usuario',
            'Sucursal'
        ]

        # Escribir encabezados en la primera fila
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos de adherentes en filas
        for row_num, adherente in enumerate(adherentes, start=2):
            ws.cell(row=row_num, column=1, value=adherente.adherente_id)
            ws.cell(row=row_num, column=2, value=adherente.titular.name)
            ws.cell(row=row_num, column=3, value=adherente.titular.last_name)
            ws.cell(row=row_num, column=4, value=adherente.titular.document_type)
            ws.cell(row=row_num, column=5, value=adherente.titular.document)
            ws.cell(row=row_num, column=6, value=adherente.titular.cbu)
            ws.cell(row=row_num, column=7, value=adherente.name)
            ws.cell(row=row_num, column=8, value=adherente.last_name)
            ws.cell(row=row_num, column=9, value=adherente.phone)
            address_complete = str(f"{adherente.street_address} {adherente.number} {adherente.floor} - {adherente.city} - {adherente.province}")
            ws.cell(row=row_num, column=10, value=address_complete)
            ws.cell(row=row_num, column=11, value=adherente.document_type)
            ws.cell(row=row_num, column=12, value=adherente.document)
            ws.cell(row=row_num, column=13, value=adherente.created.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=14, value=adherente.user_upload.username)
            ws.cell(row=row_num, column=15, value=adherente.sucursal)

        # Crear el archivo Excel en memoria
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="informe.xlsx"'
        # Guardar el libro de trabajo en la respuesta HTTP
        wb.save(response)
        return response
    return render(request, 'estadisticas.html')

@permission_required('pianoadherentes.can_view_stats', raise_exception=True)
def usuario_log(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('init_date')
        end_date_str = request.POST.get('end_date')
        usuario = request.POST.get('usuario')

        # Convertir las fechas de cadena a objetos datetime
        start_date_parts = start_date_str.split('-')
        start_date = '-'.join(start_date_parts)

        end_date_parts = end_date_str.split('-')
        end_date = '-'.join(end_date_parts)

        # Filtrar adherentes dentro del rango de fechas

        if usuario == None:
            log = Log.objects.filter(
                created__range=(start_date, end_date))
        else:
            log = Log.objects.filter(
                created__range=(start_date, end_date),
                user=usuario)

        # Crear un archivo Excel utilizando openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = 'Log por usuarios'

        # Definir encabezados de columnas
        headers = [
            'Log ID',
            'Adherente',
            'Movimiento',
            'Usuario',
            'Historia',
            'Fecha'
        ]

        # Escribir encabezados en la primera fila
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos de adherentes en filas
        for row_num, log in enumerate(log, start=2):
            ws.cell(row=row_num, column=1, value=log.log_id)
            adherente = str(f'{log.adherente.name} {log.adherente.last_name}')
            ws.cell(row=row_num, column=2, value=adherente)
            ws.cell(row=row_num, column=3, value=log.movimiento)
            ws.cell(row=row_num, column=4, value=log.user.username)
            ws.cell(row=row_num, column=5, value=log.historia)
            ws.cell(row=row_num, column=6, value=log.created.strftime('%Y-%m-%d %H:%M:%S'))
            

        # Crear el archivo Excel en memoria
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="informe.xlsx"'
        # Guardar el libro de trabajo en la respuesta HTTP
        wb.save(response)
        return response
    return render(request, 'estadisticas.html')


@permission_required('pianoadherentes.can_view_stats_controllerAdmin', raise_exception=True)
def padron_activo(request):

    if request.method == 'POST':
        adherentes = Adherente.objects.filter(is_active=1)
        # Crear un archivo Excel utilizando openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = 'Adherentes'
        # Definir encabezados de columnas
        headers = [
            'Adherente ID',
            'Titular Nombre',
            'Titular Apellido',
            'Tipo de documento',
            'Documento',
            'CBU Titular',
            'Nombre',
            'Apellido',
            'Teléfono',
            'Dirección',
            'Tipo documento',
            'Nro documento',
            'Fecha de Creación',
            'Usuario',
            'Sucursal'
        ]

        # Escribir encabezados en la primera fila
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos de adherentes en filas
        for row_num, adherente in enumerate(adherentes, start=2):
            ws.cell(row=row_num, column=1, value=adherente.adherente_id)
            ws.cell(row=row_num, column=2, value=adherente.titular.name)
            ws.cell(row=row_num, column=3, value=adherente.titular.last_name)
            ws.cell(row=row_num, column=4, value=adherente.titular.document_type)
            ws.cell(row=row_num, column=5, value=adherente.titular.document)
            ws.cell(row=row_num, column=6, value=adherente.titular.cbu)
            ws.cell(row=row_num, column=7, value=adherente.name)
            ws.cell(row=row_num, column=8, value=adherente.last_name)
            ws.cell(row=row_num, column=9, value=adherente.phone)
            address_complete = str(f"{adherente.street_address} {adherente.number} {adherente.floor} - {adherente.city} - {adherente.province}")
            ws.cell(row=row_num, column=10, value=address_complete)
            ws.cell(row=row_num, column=11, value=adherente.document_type)
            ws.cell(row=row_num, column=12, value=adherente.document)
            if adherente.created:
                ws.cell(row=row_num, column=13, value=adherente.created.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_num, column=13, value='N/A') 
            ws.cell(row=row_num, column=14, value=adherente.user_upload.username)
            ws.cell(row=row_num, column=15, value=adherente.sucursal)

        # Crear el archivo Excel en memoria
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="informe.xlsx"'
        # Guardar el libro de trabajo en la respuesta HTTP
        wb.save(response)
        return response
    return render(request, 'estadisticas.html')
